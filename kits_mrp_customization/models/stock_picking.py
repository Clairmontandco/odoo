from odoo import _, api, fields, models, tools
from io import BytesIO
import base64
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Border, Side, Alignment, Font

class StockPicking(models.Model):
    _inherit = 'stock.picking'

    invoice_payment_status = fields.Selection(related = 'sale_id.invoice_payment_status')
    paid_date = fields.Date('Paid Date',related='sale_id.paid_date',store=True)
    sale_user_id = fields.Many2one('res.users',related = 'sale_id.user_id')
    sale_team_id = fields.Many2one('crm.team',related = 'sale_id.team_id')
    payment_term_id = fields.Many2one('account.payment.term',related = 'sale_id.payment_term_id',store=True)
    backorder_count = fields.Integer('Back Order Picking',compute = '_compute_backorder_count')
    book_date =  fields.Datetime('Book date',related = 'sale_id.book_date')

    def write(self, values):
        for rec in self:
            res = super(StockPicking,self).write(values)
            invoice = rec.sale_id.invoice_ids.filtered(lambda x:x.state != 'cancel')
            if values.get('state'):
                invoice.delivery_status = rec.state
            if values.get('date_done'):
                invoice.date_done = rec.date_done
            return res  

    def action_kits_batch_order_excel(self):
        wb = Workbook()
        wrksheet = wb.active
        wrksheet.title = "Batch Order Excel"        
        header_font = Font(name='Lato', size=12, bold=True)
        value_font = Font(name='Lato', size=12, bold=False)
        
        wrksheet.cell(row=1,column=1).value = "Scheduled Date"
        wrksheet.cell(row=1,column=1).font = header_font
        wrksheet.column_dimensions['A'].width = 20
        wrksheet.cell(row=1,column=2).value = "Deadline"
        wrksheet.cell(row=1,column=2).font = header_font
        wrksheet.column_dimensions['B'].width = 20
        
        wrksheet.cell(row=1,column=3).value = "Contact"
        wrksheet.cell(row=1,column=3).font = header_font
        wrksheet.column_dimensions['C'].width = 30
        wrksheet.cell(row=1,column=4).value = "Source Document"
        wrksheet.cell(row=1,column=4).font = header_font
        wrksheet.column_dimensions['D'].width = 20
        
        wrksheet.cell(row=1,column=5).value = "Product"
        wrksheet.cell(row=1,column=5).font = header_font
        wrksheet.column_dimensions['E'].width = 30
        wrksheet.cell(row=1,column=6).value = "Internal Reference"
        wrksheet.cell(row=1,column=6).font = header_font
        wrksheet.column_dimensions['F'].width = 20
        wrksheet.cell(row=1,column=7).value = "Product Category"
        wrksheet.cell(row=1,column=7).font = header_font
        wrksheet.column_dimensions['G'].width = 30
        wrksheet.cell(row=1,column=8).value = "Demand"
        wrksheet.cell(row=1,column=8).font = header_font
        wrksheet.column_dimensions['H'].width = 10
        row_index = 2
        for rec in self.filtered(lambda x:x.state != 'cancel'):
            
            wrksheet.cell(row=row_index, column=1).value = rec.scheduled_date
            wrksheet.cell(row=row_index, column=2).value = rec.date_deadline
            wrksheet.cell(row=row_index, column=3).value = rec.partner_id.name
            wrksheet.cell(row=row_index, column=4).value = rec.origin
            wrksheet.cell(row=row_index, column=1).font = value_font
            wrksheet.cell(row=row_index, column=2).font = value_font
            wrksheet.cell(row=row_index, column=3).font = value_font
            wrksheet.cell(row=row_index, column=4).font = value_font
        
            move_ids = rec.move_ids_without_package.filtered(lambda x:x.state != 'cancel')
            for move in move_ids:
                # Value
                wrksheet.cell(row=row_index, column=5).value = move.name
                wrksheet.cell(row=row_index, column=6).value = move.product_id.default_code
                wrksheet.cell(row=row_index, column=7).value = move.product_id.categ_id.complete_name
                wrksheet.cell(row=row_index, column=8).value = move.product_uom_qty
               
                
                # Adding Font
                wrksheet.cell(row=row_index, column=5).font = value_font
                wrksheet.cell(row=row_index, column=6).font = value_font
                wrksheet.cell(row=row_index, column=7).font = value_font
                wrksheet.cell(row=row_index, column=8).font = value_font
               
                
                row_index += 1
                    
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wizard_id = self.env['report.wizard'].create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wizard_id.id,'Batch_Order_Excel'),
            'target': 'self',
        }

    @api.depends('backorder_ids')
    def _compute_backorder_count(self):
        for rec in self:
            rec.backorder_count = len(rec.backorder_ids)

    def preview_sale_order(self):
        for rec in self:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'sale.order',
                'view_type': 'form',
                'view_mode': 'form',
                'res_id': rec.sale_id.id,
                'target': 'current'
            }

    def action_view_stock_picking(self):
        self.ensure_one()
        action =  {
            'name':_('Back Order Delivery'),
            'type':'ir.actions.act_window',
            'res_model':"stock.picking",
            'target':'current',
        }
        if len(self.backorder_ids) == 1:
            action.update({
                'res_id':self.backorder_ids.id,
                'view_mode':'form',
            })
        else:
            action.update({
                'view_mode':'tree,form',
                'domain':[('id','in',self.backorder_ids.ids)],
                })
        return action
