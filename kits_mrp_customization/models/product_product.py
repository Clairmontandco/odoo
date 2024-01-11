from odoo import _,models,fields,api
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
import calendar
import base64
from odoo.exceptions import UserError

class ProductProduct(models.Model):
    _inherit = 'product.product'

    # Method for monthly data of current and previws year.
    def get_report_data(self):
        product_report_data = {}
        range_of_year = (datetime.now().year,datetime.now().year-1,datetime.now().year-2)
        months_of_year = list(calendar.month_name)

        for year in range_of_year:
            for rec in self:
               
                quary = ''' 
                            SELECT sol.id FROM sale_order_line AS sol
                            WHERE product_id = %s and (SELECT EXTRACT(YEAR FROM sol.create_date)) = %s 
                        ''' % (rec.id, year)
                self._cr.execute(quary)
                sale_order_line_ids = self._cr.fetchall()

                if not rec.name in product_report_data.keys():
                    product_report_data[rec.name] = {
                        'product_categ': rec.categ_id.display_name,
                        'product_id' : rec.id,
                        'sale_price': rec.lst_price or 0.0,
                        'stock_or_not': 'True' if rec.x_studio_stocked else 'False',
                        'cost': rec.standard_price or 0.0,
                        'profit_margin': rec.lst_price / rec.standard_price if rec.standard_price else rec.lst_price,
                        'create_date': rec.create_date.date() or '',
                    }

                sale_order_line_ids = self.env['sale.order.line'].browse([id[0] for id in sale_order_line_ids])
                
                # Prepared dict for month wize data.
                for month in months_of_year:
                    if month:
                        month_order_lines = sale_order_line_ids.filtered(lambda x: x.create_date.month == months_of_year.index(month))
                        avg_sale_price = sum(month_order_lines.mapped('price_unit'))/len(month_order_lines.mapped('price_unit')) if month_order_lines else 0.0
                        product_qty = sum(month_order_lines.mapped('product_uom_qty'))
                        if year not in product_report_data.get(rec.name).keys():
                            product_report_data.get(rec.name).update(
                                {
                                    year: {
                                        month: {
                                            'sale_price': avg_sale_price or 0.0,
                                            'product_qty': product_qty,
                                            'total': avg_sale_price * product_qty
                                        }
                                    }
                                }
                            )
                        else:
                            if month in product_report_data.get(rec.name).get(year).keys():
                                product_report_data.get(rec.name).get(year).get(month).update()
                            else:
                                product_report_data.get(rec.name).get(year).update(
                                    {
                                        month: {
                                            'sale_price': avg_sale_price or 0.0,
                                            'product_qty': product_qty,
                                            'total': avg_sale_price * product_qty
                                        }
                                    }
                                )
        return product_report_data
    
    # Method for monthly data of product consume at manufacturing.
    def get_yearly_product_consume_table_data(self):
        query = """CREATE OR REPLACE FUNCTION GET_CONSUME_PRODUCT_DATA_SPT(ARR INTEGER[], COMPANYID INTEGER) 
                    RETURNS JSON[] AS $$
                    DECLARE
                            min_year INTEGER;
                            max_year INTEGER;
                            year_diff INTEGER;
                            y INTEGER;
                            mon INTEGER;
                            product INTEGER;
                            productQty REAL;
                            productCost REAL;
                            product_data_arr JSON[];
                            product_data JSON[];
                            product_year_data JSON[];
                            resourse_name TEXT;
                            standard_price REAL;
                            consume_cost REAL;
                    BEGIN
                        IF arr IS NOT null THEN
                            DROP TABLE IF EXISTS ConsumeTempTable;
                            CREATE TEMPORARY TABLE ConsumeTempTable AS
                            (SELECT sol.id,
                                    sol.create_date,
                                    sol.qty_delivered AS LineDelivered,
                                    mbl.product_qty AS Qty,
                                    pp.id AS ProductId,
                                    mb.product_tmpl_id AS ProductTmplId,
                                    mb.id AS BomId,
                                    mbl.product_id AS SelectedProduct
                                    FROM mrp_bom_line mbl
                            LEFT JOIN mrp_bom mb ON mbl.bom_id = mb.id
                            LEFT JOIN product_product pp ON mb.product_tmpl_id = pp.product_tmpl_id
                            LEFT JOIN sale_order_line sol ON pp.id = sol.product_id
                            WHERE mbl.product_id = ANY (arr)
                                AND sol.qty_delivered > 0);
                            END IF;
                            min_year := (SELECT min(EXTRACT(year FROM ctt.create_date)) FROM ConsumeTempTable ctt);
                            max_year := (SELECT max(EXTRACT(year FROM ctt.create_date)) FROM ConsumeTempTable ctt);
                            year_diff := max_year - min_year;
                            IF year_diff < 2 THEN
                                IF year_diff = 0 THEN
                                    min_year := min_year - 2 ;
                                END IF;
                                IF year_diff = 1 THEN
                                    min_year := min_year - 1 ;
                                END IF;
                            END IF;
                            product_data_arr := json_build_object('min_year',min_year,'max_year',max_year)||product_data_arr;
                            FOREACH product IN ARRAY arr LOOP
                                resourse_name := (SELECT 'product.product,' || product);
                                standard_price := (SELECT COALESCE(value_float,0.00) FROM ir_property WHERE NAME = 'standard_price' AND res_id = resourse_name AND company_id = CompanyId);
                                product_year_data := json_build_object();
                                FOR y IN SELECT generate_series(max_year, min_year, -1) LOOP
                                product_data := json_build_object();
                                    FOR mon IN SELECT generate_series(12, 1, -1) LOOP
                                        SELECT COALESCE(SUM(ConsumeQty), 0) AS TotalConsumeQty
                                        INTO productQty
                                        FROM (
                                            SELECT COALESCE((ctt.LineDelivered * ctt.Qty),0) AS ConsumeQty
                                            FROM ConsumeTempTable ctt
                                            WHERE EXTRACT(year FROM ctt.create_date) = y AND EXTRACT(month FROM ctt.create_date) = mon AND ctt.SelectedProduct = product
                                        ) AS FinalTotal;
                                        consume_cost := productQty * standard_price;
                                    product_data = json_build_object(mon::TEXT, json_build_object('Conqty', productQty,'Conprice',consume_cost)) || product_data;
                                    END LOOP;
                                    product_year_data = json_build_object(y::TEXT,product_data) || product_year_data;
                                END LOOP;
                                product_data_arr := product_data_arr || json_build_object(product::TEXT, product_year_data);
                            END LOOP;
                        RETURN product_data_arr::json[];
                    END;
                    $$ LANGUAGE PLPGSQL;
                    SELECT GET_CONSUME_PRODUCT_DATA_SPT(ARRAY%s,%s)"""%(self.ids,self.env.company.id)
        self._cr.execute(query)
        data = self._cr.fetchall()
        if len(data) >= 1 and data and len(data[0][0]) >= 2:
            return data[0][0]
        else:
            return {}
                    
    # Method for get yearly data.
    def get_yearly_table_data(self):
        query = """CREATE OR REPLACE FUNCTION get_product_data_spt(arr INTEGER[])
                    RETURNS JSON[] AS
                    $$
                    DECLARE
                        min_year INTEGER;
                        max_year INTEGER;
                        y INTEGER;
                        product INTEGER;
                        yearly_data JSON;
                        product_data_arr JSON[];
                        product_data json;
                        product_yr_data json;
                    BEGIN
                        min_year := (select min(EXTRACT(year from sol.create_date)) from sale_order_line as sol
                                        left join product_product as pp on sol.product_id = pp.id
                                        where pp.id = any (arr));
                                        
                        max_year := (select max(EXTRACT(year from sol.create_date)) from sale_order_line as sol
                                        left join product_product as pp on sol.product_id = pp.id
                                        where pp.id = any (arr));

                        product_data_arr := ARRAY[(json_build_object('min',min_year )),(json_build_object('max',max_year ))];
						SELECT ARRAY_AGG(product_id ORDER BY TOTAL_ORDERED DESC) AS product_ids
						FROM (
							SELECT DISTINCT product_id,
								   ROUND(SUM(PRODUCT_UOM_QTY), 2) AS TOTAL_ORDERED
							FROM sale_order_line
							WHERE date_part('year', create_date) = (SELECT date_part('year', (SELECT current_timestamp)))
								  AND product_id = any(arr)
							GROUP BY product_id
						) subquery into arr;
                        if arr is not null then
                            FOREACH product IN ARRAY arr
                            LOOP
                                yearly_data := '{}';
                                FOR Y IN SELECT generate_series(max_year, min_year, -1)
                                    LOOP	
                                        product_yr_data := json_build_object(y,(SELECT ROW_TO_JSON(X) FROM (SELECT 
                                                ROUND(AVG(PRICE_UNIT),2) AS AVERAGE_PRICE,
                                                ROUND(SUM(PRODUCT_UOM_QTY),2) AS TOTAL_ORDERED,
                                                ROUND(AVG(PRICE_UNIT) * SUM(PRODUCT_UOM_QTY),2) AS TOTAL
                                            FROM SALE_ORDER_LINE AS SOL
                                            LEFT JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                                            WHERE EXTRACT(YEAR FROM SOL.CREATE_DATE) = y AND PRODUCT_ID = product GROUP BY EXTRACT(YEAR FROM SOL.CREATE_DATE))X));				

                                        yearly_data := yearly_data ::JSONB || (json_build_object(Y, (product_yr_data->>Y::TEXT)))::JSONB;
                                    END LOOP;
                                product_data_arr := product_data_arr || json_build_object(product,yearly_data);
                            END LOOP;
                        end if;
                        
                        RETURN product_data_arr::json[];
                    END;
                    $$
                    LANGUAGE plpgsql;
                    select get_product_data_spt(ARRAY%s)"""%(self.ids)
        self._cr.execute(query)
        data = self._cr.fetchall()
        if len(data) >= 1 and data and len(data[0][0]) >= 3:
            return data[0][0]
        else:
            return {}

    # Method for Monthly Sales Report
    def action_create_monthly_sales_report(self):

        report_data = self.get_report_data()
        yearly_table_data = self.get_yearly_table_data()
        if not yearly_table_data:
            raise UserError('There are no yearly data available')
        f_name = 'Sales_Spreadsheet'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Sales Forecast", index=0)
        center_aligment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        left_aligment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        right_aligment = styles.Alignment(horizontal="right", vertical="center",wrap_text=True)
        bd = styles.Side(style='thin', color="FFFFFF")
        bottom_border = styles.Border(bottom=bd)
        top_border = styles.Border(top=bd)
        top_bottom_border = styles.Border(top=bd, bottom=bd)
        header_row = 2

        sheet.row_dimensions[1].height = 12

        # Sheet header
        sheet.merge_cells('B2:F3')
        sheet.cell(row=header_row, column=2).value = "SALES FORECAST"
        sheet.cell(row=header_row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
        sheet.cell(row=header_row, column=2).font = styles.Font(name='Arial',bold=True, size=36,color='4472c4')
        sheet.row_dimensions[header_row].height = 21
        sheet.row_dimensions[header_row+1].height = 21

        sheet.row_dimensions[header_row+2].height = 10

        # Table header
        table_header_row = header_row + 3
        sheet.cell(row=table_header_row, column=2).value = "PRODUCT NAME"
        sheet.cell(row=table_header_row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
        sheet.cell(row=table_header_row, column=2).font = styles.Font(name='Arial',bold=True, size=12,color='ffffff')
        sheet.row_dimensions[table_header_row].height = 24

        # Set row bg color.
        for row in range(table_header_row,table_header_row+1):
            for col in range(2,19):
                sheet.cell(row=row, column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
        
        # For yearly table header.
        year_table_header_row = table_header_row
        for y in (yearly_table_data[1].get('max'),yearly_table_data[0].get('min')):
            for row in range(year_table_header_row,year_table_header_row+1):
                for col in range(20,20+len(list(range(yearly_table_data[1].get('max'), yearly_table_data[0].get('min') - 1, -1)))):
                    sheet.cell(row=row,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
        
        # Table content.
        table_content_row = table_header_row + 1
        content_col = 2
        next_product_index = 2
        month_total_dict = {}
        year_total_dict = {}
        for data in yearly_table_data[2:]:
            
            # Product name.
            # product_id = self.browse(report_data.get(data).get('product_id'))
            product_id = self.browse(eval(list(data.keys())[0]))
            data = product_id.name

            sheet.merge_cells('B%s:B%s'%(table_content_row,table_content_row+11))
            sheet.cell(row=table_content_row,column=content_col).value = product_id.display_name
            sheet.cell(row=table_content_row,column=content_col).alignment = center_aligment
            sheet.cell(row=table_content_row+12,column=2).fill = styles.PatternFill("solid",start_color="7f7f7f")
            content_col += 1
            
            # Product category.
            sheet.merge_cells('C%s:C%s'%(table_content_row,table_content_row+6))
            sheet.cell(row=table_content_row,column=content_col).value = report_data.get(data).get('product_categ')
            sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+12,column=3).fill = styles.PatternFill("solid",start_color="7f7f7f")

            product_tags = ''
            if product_id.product_tmpl_id.x_studio_many2many_field_bOjgj:
                product_tags = ', '.join(product_id.product_tmpl_id.x_studio_many2many_field_bOjgj.mapped('display_name')) if product_id.product_tmpl_id.x_studio_many2many_field_bOjgj else ''

            # Product tags.
            sheet.merge_cells('C%s:C%s'%(table_content_row+7,table_content_row+11))
            sheet.cell(row=table_content_row+7,column=content_col).value = 'Tags : ' + product_tags
            sheet.cell(row=table_content_row+7,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+12,column=3).fill = styles.PatternFill("solid",start_color="7f7f7f")
            content_col += 1
            
            # Print monthly data in first table (Jan - Dec)
            month_row = table_content_row + 1
            for year in (datetime.now().year-2,datetime.now().year-1,datetime.now().year):
                col = 6
                total_sale_price = 0.0
                total_product_qty = 0.0
                month_total = 0.0
                sheet.cell(row=month_row-1,column=content_col+1).fill = styles.PatternFill("solid",start_color="7f7f7f")
                sheet.cell(row=month_row,column=content_col+1).value = "SALE PRICE"
                sheet.cell(row=month_row,column=content_col+1).fill = styles.PatternFill("solid",start_color="d8d8d8")
                sheet.cell(row=month_row+1,column=content_col+1).value = "UNITS SOLD"
                sheet.cell(row=month_row+1,column=content_col+1).fill = styles.PatternFill("solid",start_color="f2f2f2")
                sheet.cell(row=month_row+2,column=content_col+1).value = "TOTAL"
                sheet.cell(row=month_row+2,column=content_col+1).font = styles.Font(name='Arial',bold=True)
                sheet.cell(row=month_row+2,column=content_col+1).fill = styles.PatternFill("solid",start_color="d8d8d8")
                sheet.cell(row=month_row+3,column=content_col+1).fill = styles.PatternFill("solid",start_color="7f7f7f")

                # Loop of year months.
                for month_data in report_data.get(data).get(year):

                    month_short_name = month_data[:3] + '-' + str(year)[-2:]

                    sheet.cell(row=month_row-1,column=col).value = month_short_name
                    sheet.cell(row=month_row-1,column=col).alignment = center_aligment
                    sheet.cell(row=month_row-1,column=col).fill = styles.PatternFill("solid",start_color="bf9000")
                    sheet.cell(row=month_row-1,column=col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)

                    sheet.cell(row=month_row,column=col).value = '$ {:,.2f}'.format(report_data.get(data).get(year).get(month_data).get('sale_price'),2)
                    sheet.cell(row=month_row,column=col).alignment = right_aligment
                    sheet.cell(row=month_row,column=col).fill = styles.PatternFill("solid",start_color="ffe598")
                    total_sale_price += report_data.get(data).get(year).get(month_data).get('sale_price')

                    sheet.cell(row=month_row+1,column=col).value = report_data.get(data).get(year).get(month_data).get('product_qty')
                    sheet.cell(row=month_row+1,column=col).alignment = right_aligment
                    sheet.cell(row=month_row+1,column=col).fill = styles.PatternFill("solid",start_color="fef2cb")
                    total_product_qty += report_data.get(data).get(year).get(month_data).get('product_qty')

                    sheet.cell(row=month_row+2,column=col).value = '$ {:,.2f}'.format(report_data.get(data).get(year).get(month_data).get('total'))
                    sheet.cell(row=month_row+2,column=col).font = styles.Font(name='Arial',bold=True)
                    sheet.cell(row=month_row+2,column=col).fill = styles.PatternFill("solid",start_color="ffe598")
                    month_total += report_data.get(data).get(year).get(month_data).get('total')

                    sheet.cell(row=month_row+2,column=col).alignment = right_aligment

                    sheet.cell(row=month_row+3,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")

                    # Prepared data for month total.
                    if year in month_total_dict.keys():
                        if not month_data in month_total_dict.get(year).keys():
                            month_total_dict.get(year)[month_data] = {
                                'total_qty':report_data.get(data).get(year).get(month_data).get('product_qty'),
                                'total':report_data.get(data).get(year).get(month_data).get('total')
                            }
                        else:
                            month_total_dict.get(year).get(month_data).update({
                                'total_qty':month_total_dict.get(year).get(month_data).get('total_qty') + report_data.get(data).get(year).get(month_data).get('product_qty'),
                                'total':month_total_dict.get(year).get(month_data).get('total') + report_data.get(data).get(year).get(month_data).get('total')
                            })
                    else:
                        month_total_dict[year] = {
                            month_data : {
                                'total_qty': report_data.get(data).get(year).get(month_data).get('product_qty'),
                                'total':report_data.get(data).get(year).get(month_data).get('total')
                            }
                        }
                    col += 1

                # Month total section.
                sheet.cell(row=month_row-1,column=col).value = 'Total'
                sheet.cell(row=month_row-1,column=col).alignment = center_aligment
                sheet.cell(row=month_row-1,column=col).fill = styles.PatternFill("solid",start_color="2f5496")
                sheet.cell(row=month_row-1,column=col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)

                sheet.cell(row=month_row,column=col).value = '$ {:,.2f}'.format(total_sale_price)
                sheet.cell(row=month_row,column=col).alignment = right_aligment
                sheet.cell(row=month_row,column=col).fill = styles.PatternFill("solid",start_color="b4c6e7")

                sheet.cell(row=month_row+1,column=col).value = total_product_qty
                sheet.cell(row=month_row+1,column=col).alignment = right_aligment
                sheet.cell(row=month_row+1,column=col).fill = styles.PatternFill("solid",start_color="d9e2f3")

                sheet.cell(row=month_row+2,column=col).value = '$ {:,.2f}'.format(month_total)
                sheet.cell(row=month_row+2,column=col).alignment = right_aligment
                sheet.cell(row=month_row+2,column=col).fill = styles.PatternFill("solid",start_color="b4c6e7")

                sheet.cell(row=month_row+3,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")

                month_row += 4
            sheet.cell(row=table_content_row,column=content_col).border = bottom_border

            # Product information section.
            sheet.cell(row=table_content_row+1,column=content_col).value = "Sale Price : "+'$ {:,.2f}'.format(report_data.get(data).get('sale_price')) or 0.0
            sheet.cell(row=table_content_row+1,column=content_col).border = bottom_border
            sheet.cell(row=table_content_row+1,column=content_col).alignment = left_aligment
            table_content_row += 2

            sheet.cell(row=table_content_row+1,column=content_col).value = "Qty On Hand : " + str(round(product_id.qty_available,2)) or 0
            sheet.cell(row=table_content_row+1,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+1,column=content_col).border = top_bottom_border
            table_content_row += 2

            sheet.cell(row=table_content_row+1,column=content_col).value = "In Stock : " + report_data.get(data).get('stock_or_not')
            sheet.cell(row=table_content_row+1,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+1,column=content_col).border = top_bottom_border
            table_content_row += 2

            sheet.cell(row=table_content_row+1,column=content_col).value = "Cost : " + '$ {:,.2f}'.format(report_data.get(data).get('cost')) or 0.0
            sheet.cell(row=table_content_row+1,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+1,column=content_col).border = top_bottom_border
            table_content_row += 1

            sheet.cell(row=table_content_row+1,column=content_col).value = "Margin : " + '$ {:,.2f}'.format(report_data.get(data).get('sale_price') - report_data.get(data).get('cost')) or 0.0
            sheet.cell(row=table_content_row+1,column=content_col).alignment = left_aligment
            sheet.cell(row=table_content_row+1,column=content_col).border = top_bottom_border
            table_content_row += 2
            
            key_web_accounts = ''
            # If product key/web accounts then merge rows.
            if product_id.product_tmpl_id.x_studio_many2many_field_T5tHX:
                key_web_accounts = ', '.join(product_id.product_tmpl_id.x_studio_many2many_field_T5tHX.mapped('display_name')) if product_id.product_tmpl_id.x_studio_many2many_field_T5tHX else ''
            

            sheet.cell(row=table_content_row+1,column=content_col).value = "Create Date : " + str(report_data.get(data).get('create_date'))
            sheet.cell(row=table_content_row+1,column=content_col).alignment = styles.Alignment(horizontal="left", vertical="top",wrap_text=True)
            sheet.cell(row=table_content_row+1,column=content_col).border = top_border

            sheet.cell(row=table_content_row+2,column=content_col).value = "Key/Web Accounts : " + key_web_accounts
            sheet.cell(row=table_content_row+2,column=content_col).alignment = styles.Alignment(horizontal="left", vertical="top",wrap_text=True)
            sheet.cell(row=table_content_row+2,column=content_col).border = top_border
            table_content_row += 4

            # Row seperator with color.
            sheet.cell(row=table_content_row-1,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.row_dimensions[table_content_row-1].height = 10
            # table_content_row += 2

            year_table_col = col + 2
            product_data = yearly_table_data[next_product_index] or False
            product_name_row = table_header_row+2
            year_range = list(range(yearly_table_data[1].get('max'), yearly_table_data[0].get('min') - 1, -1))

            # Merge yearly data table header.
            merge_col_range = get_column_letter(20+len(list(range(yearly_table_data[1].get('max'), yearly_table_data[0].get('min'), -1))))
            sheet.merge_cells('T%s:%s%s'%(table_header_row+2,merge_col_range,table_header_row+2))
            sheet.cell(row=product_name_row,column=year_table_col).value = product_id.display_name
            sheet.cell(row=product_name_row,column=year_table_col).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)

            # Add value of yearly data in table.
            for year in list(range(yearly_table_data[1].get('max'), yearly_table_data[0].get('min') - 1, -1)):
                converted_dict_data = (product_data.get(str(product_id.id),{}).get(str(year))) if product_data.get(str(product_id.id),{}).get(str(year)) else {}
                if isinstance(converted_dict_data,str):
                    converted_dict_data = eval(converted_dict_data)
                else:
                    converted_dict_data = {}
                if product_data and product_data.get(str(product_id.id),{}).get(str(year)):

                    sheet.cell(row=table_header_row+1,column=year_table_col).value = year
                    sheet.cell(row=table_header_row+1,column=year_table_col).fill = styles.PatternFill("solid",start_color="bf9000")
                    sheet.cell(row=table_header_row+1,column=year_table_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                    sheet.cell(row=table_header_row+1,column=year_table_col).alignment = center_aligment

                    sheet.cell(row=table_header_row+3,column=year_table_col).value = '$ {:,.2f}'.format((converted_dict_data).get('average_price',0))
                    sheet.cell(row=table_header_row+3,column=year_table_col).alignment = right_aligment
                    sheet.cell(row=table_header_row+3,column=year_table_col).fill = styles.PatternFill("solid",start_color="ffe598")

                    sheet.cell(row=table_header_row+4,column=year_table_col).value = converted_dict_data.get('total_ordered',0) 
                    sheet.cell(row=table_header_row+4,column=year_table_col).fill = styles.PatternFill("solid",start_color="fef2cb")

                    sheet.cell(row=table_header_row+5,column=year_table_col).value = '$ {:,.2f}'.format((converted_dict_data).get('total',0))
                    sheet.cell(row=table_header_row+5,column=year_table_col).alignment = right_aligment
                    sheet.cell(row=table_header_row+5,column=year_table_col).font = styles.Font(name='Arial',bold=True)
                    sheet.cell(row=table_header_row+5,column=year_table_col).fill = styles.PatternFill("solid",start_color="bf9000")
                else:
                    sheet.cell(row=table_header_row+1,column=year_table_col).value = year
                    sheet.cell(row=table_header_row+1,column=year_table_col).fill = styles.PatternFill("solid",start_color="bf9000")
                    sheet.cell(row=table_header_row+1,column=year_table_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                    sheet.cell(row=table_header_row+1,column=year_table_col).alignment = center_aligment

                    sheet.cell(row=table_header_row+3,column=year_table_col).value = '$ {:,.2f}'.format(0.0)
                    sheet.cell(row=table_header_row+3,column=year_table_col).alignment = right_aligment
                    sheet.cell(row=table_header_row+3,column=year_table_col).fill = styles.PatternFill("solid",start_color="ffe598")

                    sheet.cell(row=table_header_row+4,column=year_table_col).value = 0
                    sheet.cell(row=table_header_row+4,column=year_table_col).fill = styles.PatternFill("solid",start_color="fef2cb")

                    sheet.cell(row=table_header_row+5,column=year_table_col).value = '$ {:,.2f}'.format(0.0)
                    sheet.cell(row=table_header_row+5,column=year_table_col).alignment = right_aligment
                    sheet.cell(row=table_header_row+5,column=year_table_col).font = styles.Font(name='Arial',bold=True)
                    sheet.cell(row=table_header_row+5,column=year_table_col).fill = styles.PatternFill("solid",start_color="bf9000")
                product_name_row += 11

                # Prepared yearly table total.
                if not year in year_total_dict.keys():
                    year_total_dict[year] = {
                        'total_qty':(converted_dict_data).get('total_ordered',0),
                        'total':(converted_dict_data).get('total',0)
                    }
                else:
                    if converted_dict_data and converted_dict_data.get('total_ordered') :
                        year_total_dict.get(year).update({'total_qty':year_total_dict.get(year).get('total_qty') + converted_dict_data.get('total_ordered',0) })
                    if converted_dict_data and converted_dict_data.get('total'):
                        year_total_dict.get(year).update({'total':year_total_dict.get(year).get('total') + converted_dict_data.get('total',0)})
                
                sheet.cell(row=table_header_row+13,column=year_table_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
                year_table_col += 1
            content_col = 2
            next_product_index += 1
            table_header_row += 13

        # Heading of monthly total.
        sheet.merge_cells('B%s:R%s'%(table_content_row,table_content_row))
        sheet.cell(row=table_content_row,column=2).value = 'MONTHLY TOTALS'
        sheet.cell(row=table_content_row,column=2).alignment = center_aligment
        sheet.cell(row=table_content_row,column=2).fill = styles.PatternFill("solid",start_color="2f5496")
        sheet.cell(row=table_content_row,column=2).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
        table_content_row += 1
        content_col += 3

        sheet.merge_cells('B%s:B%s'%(table_content_row,table_content_row+6))
        sheet.merge_cells('C%s:C%s'%(table_content_row,table_content_row+6))
        sheet.merge_cells('D%s:D%s'%(table_content_row,table_content_row+6))

        sheet.cell(row=table_content_row,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")

        sheet.cell(row=table_content_row+1,column=content_col).value = 'UNITS SOLD'
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
        table_content_row += 1

        sheet.cell(row=table_content_row+1,column=content_col).value = 'TOTAL'
        sheet.cell(row=table_content_row+1,column=content_col).font = styles.Font(name='Arial',bold=True)
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")
        table_content_row += 1

        # sheet.cell(row=table_content_row+1,column=content_col).value = 'TOTAL'
        # sheet.cell(row=table_content_row+1,column=content_col).font = styles.Font(name='Arial',bold=True)
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
        table_content_row += 1

        sheet.cell(row=table_content_row+1,column=content_col).value = 'UNITS SOLD'
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
        table_content_row += 1

        sheet.cell(row=table_content_row+1,column=content_col).value = 'TOTAL'
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")
        sheet.cell(row=table_content_row+1,column=content_col).font = styles.Font(name='Arial',bold=True,color='000000',size=12)

        #######
        table_content_row += 1
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
        table_content_row += 1

        sheet.cell(row=table_content_row+1,column=content_col).value = 'UNITS SOLD'
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
        table_content_row += 1

        sheet.cell(row=table_content_row+1,column=content_col).value = 'TOTAL'
        sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")
        sheet.cell(row=table_content_row+1,column=content_col).font = styles.Font(name='Arial',bold=True,color='000000',size=12)
        
        # Add value of monthly total.
        month_data_col = content_col+1
        total_row = table_content_row
        for y in (datetime.now().year-2,datetime.now().year-1,datetime.now().year):
            total_qty = 0.0
            main_total = 0.0
            for month_data in month_total_dict.get(y):
                sheet.cell(row=total_row-7,column=month_data_col).value = month_data[:3] + '-' + str(y)[-2:] or ''
                sheet.cell(row=total_row-7,column=month_data_col).alignment = center_aligment
                sheet.cell(row=total_row-7,column=month_data_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                sheet.cell(row=total_row-7,column=month_data_col).fill = styles.PatternFill("solid",start_color="bf9000")

                sheet.cell(row=total_row-6,column=month_data_col).value = month_total_dict.get(y).get(month_data).get('total_qty') or 0
                sheet.cell(row=total_row-6,column=month_data_col).alignment = right_aligment
                sheet.cell(row=total_row-6,column=month_data_col).font = styles.Font(name='Arial',bold=False)
                sheet.cell(row=total_row-6,column=month_data_col).fill = styles.PatternFill("solid",start_color="ffe598")

                sheet.cell(row=total_row-5,column=month_data_col).value = '$ {:,.2f}'.format(month_total_dict.get(y).get(month_data).get('total')) or 0.0
                sheet.cell(row=total_row-5,column=month_data_col).alignment = right_aligment
                sheet.cell(row=total_row-5,column=month_data_col).font = styles.Font(name='Arial',bold=True)
                sheet.cell(row=total_row-5,column=month_data_col).fill = styles.PatternFill("solid",start_color="fef2cb")
               
                total_qty += month_total_dict.get(y).get(month_data).get('total_qty')
                main_total += month_total_dict.get(y).get(month_data).get('total')
                month_data_col += 1

            sheet.cell(row=total_row-7,column=month_data_col).value = 'Total'
            sheet.cell(row=total_row-7,column=month_data_col).alignment = center_aligment
            sheet.cell(row=total_row-7,column=month_data_col).fill = styles.PatternFill("solid",start_color="2f5496")
            sheet.cell(row=total_row-7,column=month_data_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)

            sheet.cell(row=total_row-6,column=month_data_col).value = total_qty
            sheet.cell(row=total_row-6,column=month_data_col).alignment = right_aligment
            sheet.cell(row=total_row-6,column=month_data_col).fill = styles.PatternFill("solid",start_color="b4c6e7")

            sheet.cell(row=total_row-5,column=month_data_col).value = '$ {:,.2f}'.format(main_total)
            sheet.cell(row=total_row-5,column=month_data_col).alignment = right_aligment
            sheet.cell(row=total_row-5,column=month_data_col).fill = styles.PatternFill("solid",start_color="d9e2f3")

            month_data_col = content_col+1
            total_row += 3

        sheet.merge_cells('B%s:R%s'%(table_content_row+2,table_content_row+2))
        sheet.cell(row=table_content_row+2,column=2).fill = styles.PatternFill("solid",start_color="7f7f7f")
        sheet.row_dimensions[table_content_row+2].height = 10

        # Header of yearly total.
        merge_col_range = get_column_letter(20+len(year_total_dict)-1)
        sheet.merge_cells('T%s:%s%s'%(table_header_row+1,merge_col_range,table_header_row+1))
        sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).value = "YEARLY TOTALS"
        sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
        sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).fill = styles.PatternFill("solid",start_color="2f5496")
        sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).font = styles.Font(name='Arial',bold=True,color='ffffff')

        # Add value of yearly total.
        end_report_seperator = year_table_col-len(year_range)
        for r in year_range:
            year_total_data = year_total_dict.get(r)
            sheet.cell(row=table_content_row-7,column=end_report_seperator).value = year_total_data.get('total_qty') or 0.0
            sheet.cell(row=table_content_row-7,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="d9e2f3")

            sheet.cell(row=table_content_row-6,column=end_report_seperator).value = '$ {:,.2f}'.format(year_total_data.get('total') or 0.0)
            sheet.cell(row=table_content_row-6,column=end_report_seperator).font = styles.Font(name='Arial',bold=True)
            sheet.cell(row=table_content_row-6,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="b4c6e7")
            sheet.cell(row=table_content_row-6,column=end_report_seperator).alignment = right_aligment

            sheet.cell(row=table_content_row+2,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="7f7f7f")
            end_report_seperator += 1

        # Width of all col.
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 26
        sheet.column_dimensions['C'].width = 27
        sheet.column_dimensions['D'].width = 33
        sheet.column_dimensions['E'].width = 19
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 13
        sheet.column_dimensions['H'].width = 13
        sheet.column_dimensions['I'].width = 13
        sheet.column_dimensions['J'].width = 13
        sheet.column_dimensions['K'].width = 13
        sheet.column_dimensions['L'].width = 13
        sheet.column_dimensions['M'].width = 13
        sheet.column_dimensions['N'].width = 13
        sheet.column_dimensions['O'].width = 13
        sheet.column_dimensions['P'].width = 13
        sheet.column_dimensions['Q'].width = 13
        sheet.column_dimensions['R'].width = 13
        sheet.column_dimensions['S'].width = 2
        sheet.column_dimensions['T'].width = 16
        sheet.column_dimensions['U'].width = 16
        sheet.column_dimensions['V'].width = 16
        sheet.column_dimensions['W'].width = 16
        sheet.column_dimensions['X'].width = 16
        sheet.column_dimensions['Y'].width = 16
        sheet.column_dimensions['Z'].width = 16

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wizard_id = self.env['report.wizard'].create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wizard_id.id,f_name),
            'target': 'self',
        }

    def action_create_product_consume_report(self):
        months_of_year = list(calendar.month_name)
        product = self.env['product.product']
        for rec in self:
            product |= rec.filtered(lambda p:p.product_tmpl_id.bom_line_ids.ids != [])
        if product: 
            # report_data = product.get_product_consume_report_data()
            yearly_table_data = product.get_yearly_product_consume_table_data()
        f_name = 'Product_Consume_Spreadsheet'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="products consume report", index=0)
        center_aligment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        left_aligment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        right_aligment = styles.Alignment(horizontal="right", vertical="center",wrap_text=True)
        bd = styles.Side(style='thin', color="FFFFFF")
        bottom_border = styles.Border(bottom=bd)
        top_border = styles.Border(top=bd)
        top_bottom_border = styles.Border(top=bd, bottom=bd)
        header_row = 2

        sheet.row_dimensions[1].height = 12

        # Sheet header
        sheet.merge_cells('B2:F3')
        sheet.cell(row=header_row, column=2).value = "PRODUCTS CONSUME REPORT"
        sheet.cell(row=header_row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
        sheet.cell(row=header_row, column=2).font = styles.Font(name='Arial',bold=True, size=36,color='4472c4')
        sheet.row_dimensions[header_row].height = 21
        sheet.row_dimensions[header_row+1].height = 21

        sheet.row_dimensions[header_row+2].height = 10

        # Table header
        table_header_row = header_row + 3
        sheet.cell(row=table_header_row, column=2).value = "PRODUCT NAME"
        sheet.cell(row=table_header_row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
        sheet.cell(row=table_header_row, column=2).font = styles.Font(name='Arial',bold=True, size=12,color='ffffff')
        if not product:
            sheet.cell(row=table_header_row,column=2).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.cell(row=table_header_row,column=3).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.cell(row=table_header_row,column=4).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.cell(row=table_header_row,column=5).fill = styles.PatternFill("solid",start_color="7f7f7f")
        sheet.row_dimensions[table_header_row].height = 24

        # Set row bg color.
        if product:
            max_yearr = yearly_table_data[0].get('max_year')
            for row in range(table_header_row,table_header_row+1):
                for col in range(2,19):
                    sheet.cell(row=row, column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
            
            # For yearly table header.
            year_table_header_row = table_header_row
            for y in (yearly_table_data[0].get('max_year'),yearly_table_data[0].get('min_year')):
                for row in range(year_table_header_row,year_table_header_row+1):
                    for col in range(20,20+len(list(range(yearly_table_data[0].get('max_year'), yearly_table_data[0].get('min_year') - 1, -1)))):
                        sheet.cell(row=row,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
            
            # Table content.
            table_content_row = table_header_row + 1
            content_col = 2
            next_product_index = 1
            month_total_dict = {}
            year_total_dict = {}
            product_year_dict = {}
            for data in yearly_table_data[1:]:
                
                # Product name.
                product_id = self.browse(eval(list(data.keys())[0]))
                key_product = str(product_id.id)

                sheet.merge_cells('B%s:B%s'%(table_content_row,table_content_row+8))
                sheet.cell(row=table_content_row,column=content_col).value = product_id.display_name
                sheet.cell(row=table_content_row,column=content_col).alignment = center_aligment
                sheet.cell(row=table_content_row+9,column=2).fill = styles.PatternFill("solid",start_color="7f7f7f")
                content_col += 1
                
                # Product category.
                sheet.merge_cells('C%s:C%s'%(table_content_row,table_content_row+4))
                sheet.cell(row=table_content_row,column=content_col).value = product_id.categ_id.display_name
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment

                product_tags = ''
                if product_id.product_tmpl_id.x_studio_many2many_field_bOjgj:
                    product_tags = ', '.join(product_id.product_tmpl_id.x_studio_many2many_field_bOjgj.mapped('display_name')) if product_id.product_tmpl_id.x_studio_many2many_field_bOjgj else ''

                # Product tags.
                sheet.merge_cells('C%s:C%s'%(table_content_row+5,table_content_row+8))
                sheet.cell(row=table_content_row+5,column=content_col).value = 'Tags : ' + product_tags
                sheet.cell(row=table_content_row+5,column=content_col).alignment = left_aligment
                sheet.cell(row=table_content_row+9,column=3).fill = styles.PatternFill("solid",start_color="7f7f7f")
                content_col += 1
                
                # Print monthly data in first table (Jan - Dec)
                month_row = table_content_row + 1
                data_list = data.get(key_product)
                for year in (max_yearr-2,max_yearr-1,max_yearr):
                    col = 6
                    # total_sale_price = 0.0
                    total_product_qty = 0.0
                    total_product_cost = 0.0
                    sheet.cell(row=month_row-1,column=content_col+1).fill = styles.PatternFill("solid",start_color="7f7f7f")
                    sheet.cell(row=month_row+2-1-1,column=content_col+1).value = "CONSUME UNITS"
                    sheet.cell(row=month_row+2-1-1,column=content_col+1).fill = styles.PatternFill("solid",start_color="f2f2f2")
                    sheet.cell(row=month_row+3-1-1,column=content_col+1).value = "PRODUCT COST"
                    sheet.cell(row=month_row+3-1-1,column=content_col+1).fill = styles.PatternFill("solid",start_color="d8d8d8")
                    sheet.cell(row=month_row+3-1,column=content_col+1).fill = styles.PatternFill("solid",start_color="7f7f7f")

                    # Loop of year months.
                    try:
                        yearly_list = list(list(filter(lambda x : list(x.keys())[0]==str(year),data_list))[0].values())[0]
                    except Exception as ex:
                        yearly_list = []
                        raise UserError('There are no yearly data found')

                    for month_data in yearly_list:
                        month_name = months_of_year[int(list(month_data.keys())[0])]
                        month_short_name = month_name[:3] + '-' + str(year)[-2:]

                        sheet.cell(row=month_row-1,column=col).value = month_short_name
                        sheet.cell(row=month_row-1,column=col).alignment = center_aligment
                        sheet.cell(row=month_row-1,column=col).fill = styles.PatternFill("solid",start_color="77933C")
                        sheet.cell(row=month_row-1,column=col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)


                        sheet.cell(row=month_row+2-1-1,column=col).value = round((month_data.get(list(month_data.keys())[0]).get('Conqty') or 0.00),2)
                        sheet.cell(row=month_row+2-1-1,column=col).fill = styles.PatternFill("solid",start_color="c3d69b")
                        sheet.cell(row=month_row+2-1-1,column=col).alignment = right_aligment

                        sheet.cell(row=month_row+3-1-1,column=col).value = '$ {:,.2f}'.format(round((month_data.get(list(month_data.keys())[0]).get('Conprice') or 0.00),2))
                        sheet.cell(row=month_row+3-1-1,column=col).alignment = right_aligment
                        sheet.cell(row=month_row+3-1-1,column=col).fill = styles.PatternFill("solid",start_color="EBF1DE")
                        sheet.cell(row=month_row+3-1,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
                        total_product_qty += (month_data.get(list(month_data.keys())[0]).get('Conqty') or 0)
                        total_product_cost += (month_data.get(list(month_data.keys())[0]).get('Conprice') or 0)

                        # Prepared data for month total.
                        if year in month_total_dict.keys():
                            if not month_name in month_total_dict.get(year).keys():
                                month_total_dict.get(year)[month_name] = {
                                    'total_qty':(month_data.get(list(month_data.keys())[0]).get('Conqty') or 0),
                                    'total':(month_data.get(list(month_data.keys())[0]).get('Conprice') or 0)
                                }
                            else:
                                month_total_dict.get(year).get(month_name).update({
                                    'total_qty':month_total_dict.get(year).get(month_name).get('total_qty') + (month_data.get(list(month_data.keys())[0]).get('Conqty') or 0),
                                    'total':month_total_dict.get(year).get(month_name).get('total') + (month_data.get(list(month_data.keys())[0]).get('Conprice') or 0)
                                })
                        else:
                            month_total_dict[year] = {
                                month_name : {
                                    'total_qty': (month_data.get(list(month_data.keys())[0]).get('Conqty') or 0),
                                    'total':(month_data.get(list(month_data.keys())[0]).get('Conprice') or 0)
                                }
                            }
                       
                        col += 1

                    # Prepared data for year total.
                    if product_id.id in product_year_dict.keys():
                        product_year_dict[product_id.id][year] = {
                            'year_total_qty':total_product_qty,
                            'year_total':total_product_cost
                        }
                    else:
                        product_year_dict[product_id.id] = {
                            year : {
                                'year_total_qty':total_product_qty,
                                'year_total':total_product_cost
                            }
                        }
                    # Month total section.
                    sheet.cell(row=month_row-1,column=col).value = 'Total'
                    sheet.cell(row=month_row-1,column=col).alignment = center_aligment
                    sheet.cell(row=month_row-1,column=col).fill = styles.PatternFill("solid",start_color="2f5496")
                    sheet.cell(row=month_row-1,column=col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)


                    sheet.cell(row=month_row+1-1,column=col).value = round(total_product_qty,2)
                    sheet.cell(row=month_row+1-1,column=col).alignment = right_aligment
                    sheet.cell(row=month_row+1-1,column=col).fill = styles.PatternFill("solid",start_color="d9e2f3")

                    sheet.cell(row=month_row+2-1,column=col).value = '$ {:,.2f}'.format(round(total_product_cost,2))
                    sheet.cell(row=month_row+2-1,column=col).alignment = right_aligment
                    sheet.cell(row=month_row+3-1,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")
                    # sheet.cell(row=month_row+3,column=col).fill = styles.PatternFill("solid",start_color="7f7f7f")

                    month_row += 3
                sheet.cell(row=table_content_row,column=content_col).border = bottom_border

                # Product information section.
                sheet.cell(row=table_content_row,column=content_col).value = "Sale Price : "+'$ {:,.2f}'.format((product_id.lst_price or 0.0))
                sheet.cell(row=table_content_row,column=content_col).border = bottom_border
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
                table_content_row += 1

                sheet.cell(row=table_content_row,column=content_col).value = "Qty On Hand : " + str(round(product_id.qty_available,2) or 0) 
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
                sheet.cell(row=table_content_row,column=content_col).border = top_bottom_border
                table_content_row += 1

                sheet.cell(row=table_content_row,column=content_col).value = "In Stock : " + ('True' if product_id.x_studio_stocked else 'False')
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
                sheet.cell(row=table_content_row,column=content_col).border = top_bottom_border
                table_content_row += 1

                sheet.cell(row=table_content_row,column=content_col).value = "Cost : " + '$ {:,.2f}'.format((product_id.standard_price or 0.0))
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
                sheet.cell(row=table_content_row,column=content_col).border = top_bottom_border
                table_content_row += 3

                sheet.cell(row=table_content_row,column=content_col).value = "Margin : " + '$ {:,.2f}'.format((product_id.lst_price or 0.0) - (product_id.standard_price or 0.0)) 
                sheet.cell(row=table_content_row,column=content_col).alignment = left_aligment
                sheet.cell(row=table_content_row,column=content_col).border = top_bottom_border
                table_content_row += 2
                
                key_web_accounts = ''
                # If product key/web accounts then merge rows.
                if product_id.product_tmpl_id.x_studio_many2many_field_T5tHX:
                    key_web_accounts = ', '.join(product_id.product_tmpl_id.x_studio_many2many_field_T5tHX.mapped('display_name')) if product_id.product_tmpl_id.x_studio_many2many_field_T5tHX else ''
                

                sheet.cell(row=table_content_row,column=content_col).value = "Create Date : " + str(product_id.create_date.date())
                sheet.cell(row=table_content_row,column=content_col).alignment = styles.Alignment(horizontal="left", vertical="top",wrap_text=True)
                sheet.cell(row=table_content_row,column=content_col).border = top_border


                # Row seperator with color.
                sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
                sheet.row_dimensions[table_content_row+1].height = 10
                table_content_row += 2

                year_table_col = col + 2
                product_data = yearly_table_data[next_product_index] or False
                product_name_row = table_header_row+2
                year_range = list(range(yearly_table_data[0].get('max_year'), yearly_table_data[0].get('min_year') - 1, -1))

                # Merge yearly data table header.
                merge_col_range = get_column_letter(20+len(list(range(yearly_table_data[0].get('max_year'), yearly_table_data[0].get('min_year'), -1))))
                sheet.merge_cells('T%s:%s%s'%(table_header_row+2,merge_col_range,table_header_row+2))
                sheet.cell(row=product_name_row,column=year_table_col).value = product_id.display_name
                sheet.cell(row=product_name_row,column=year_table_col).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)

                # Add value of yearly data in table.
                for year in list(range(yearly_table_data[0].get('max_year'), yearly_table_data[0].get('min_year') - 1, -1)):
                    converted_dict_data = product_year_dict[(product_id.id)].get(year) if product_year_dict[(product_id.id)] else {}
                    if converted_dict_data:

                        sheet.cell(row=table_header_row+1,column=year_table_col).value = year
                        sheet.cell(row=table_header_row+1,column=year_table_col).fill = styles.PatternFill("solid",start_color="2F5496")
                        sheet.cell(row=table_header_row+1,column=year_table_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                        sheet.cell(row=table_header_row+1,column=year_table_col).alignment = center_aligment

                        sheet.cell(row=table_header_row+4-1,column=year_table_col).value = round((converted_dict_data.get('year_total_qty',0) if isinstance(converted_dict_data,dict) else 0),2)
                        sheet.cell(row=table_header_row+4-1,column=year_table_col).fill = styles.PatternFill("solid",start_color="c3d69b")
                        sheet.cell(row=table_header_row+4,column=year_table_col).value = round((converted_dict_data.get('year_total',0) if isinstance(converted_dict_data,dict) else 0),2)
                        sheet.cell(row=table_header_row+4,column=year_table_col).fill = styles.PatternFill("solid",start_color="D9E2F3")

                        sheet.cell(row=table_header_row+5-1-1,column=year_table_col).fill = styles.PatternFill("solid",start_color="B4C6E7")
                    else:
                        sheet.cell(row=table_header_row+1,column=year_table_col).value = year
                        sheet.cell(row=table_header_row+1,column=year_table_col).fill = styles.PatternFill("solid",start_color="2F5496")
                        sheet.cell(row=table_header_row+1,column=year_table_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                        sheet.cell(row=table_header_row+1,column=year_table_col).alignment = center_aligment

                        sheet.cell(row=table_header_row+4-1,column=year_table_col).value = 0
                        sheet.cell(row=table_header_row+4-1,column=year_table_col).fill = styles.PatternFill("solid",start_color="D9E2F3")

                        sheet.cell(row=table_header_row+5-1-1,column=year_table_col).fill = styles.PatternFill("solid",start_color="B4C6E7")
                    product_name_row += 10

                    # Prepared yearly table total.
                    if not year in year_total_dict.keys():
                        year_total_dict[year] = {
                            'total_qty':(converted_dict_data.get('year_total_qty',0) if isinstance(converted_dict_data,dict) else 0),
                            'total':(converted_dict_data.get('year_total',0) if isinstance(converted_dict_data,dict) else 0)
                        }
                    else:
                        if converted_dict_data and converted_dict_data.get('year_total_qty'):
                            year_total_dict.get(year).update({'total_qty':(year_total_dict.get(year).get('total_qty')or 0.00)+(converted_dict_data.get('year_total_qty',0) if isinstance(converted_dict_data,dict) else 0)})
                        if converted_dict_data and converted_dict_data.get('year_total'):
                            year_total_dict.get(year).update({'total':(year_total_dict.get(year).get('total')or 0.00)+(converted_dict_data.get('year_total',0) if isinstance(converted_dict_data,dict) else 0)})
                    
                    sheet.cell(row=table_header_row+10,column=year_table_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
                    year_table_col += 1
                content_col = 2
                next_product_index += 1
                table_header_row += 10

            # Heading of monthly total.
            sheet.merge_cells('B%s:R%s'%(table_content_row,table_content_row))
            sheet.cell(row=table_content_row,column=2).value = 'MONTHLY TOTALS'
            sheet.cell(row=table_content_row,column=2).alignment = center_aligment
            sheet.cell(row=table_content_row,column=2).fill = styles.PatternFill("solid",start_color="2f5496")
            sheet.cell(row=table_content_row,column=2).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
            table_content_row += 1
            content_col += 3

            sheet.merge_cells('B%s:B%s'%(table_content_row,table_content_row+8))
            sheet.merge_cells('C%s:C%s'%(table_content_row,table_content_row+8))
            sheet.merge_cells('D%s:D%s'%(table_content_row,table_content_row+8))

            sheet.cell(row=table_content_row,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")

            sheet.cell(row=table_content_row+1,column=content_col).value = 'CONSUME UNITS'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
            table_content_row += 1
            sheet.cell(row=table_content_row+1,column=content_col).value = 'PRODUCT COST'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")
            table_content_row += 2
            sheet.cell(row=table_content_row,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")

            sheet.cell(row=table_content_row+1,column=content_col).value = 'CONSUME UNITS'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
            table_content_row += 1
            sheet.cell(row=table_content_row+1,column=content_col).value = 'PRODUCT COST'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")

            table_content_row += 2
            sheet.cell(row=table_content_row,column=content_col).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.cell(row=table_content_row+1,column=content_col).value = 'CONSUME UNITS'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="f2f2f2")
            table_content_row += 1
            sheet.cell(row=table_content_row+1,column=content_col).value = 'PRODUCT COST'
            sheet.cell(row=table_content_row+1,column=content_col).fill = styles.PatternFill("solid",start_color="d8d8d8")

            # Add value of monthly total.
            
            month_data_col = content_col+1
            total_row = table_content_row
            for y in (max_yearr-2,max_yearr-1,max_yearr):
                total_qty = 0.0
                main_total = 0.0
                # if month_total_dict.get(y):
                for month_data in month_total_dict.get(y):
                    sheet.cell(row=total_row-7,column=month_data_col).value = month_data[:3] + '-' + str(y)[-2:] or ''
                    sheet.cell(row=total_row-7,column=month_data_col).alignment = center_aligment
                    sheet.cell(row=total_row-7,column=month_data_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)
                    sheet.cell(row=total_row-7,column=month_data_col).fill = styles.PatternFill("solid",start_color="77933C")

                    sheet.cell(row=total_row-6,column=month_data_col).value = round((month_total_dict.get(y).get(month_data).get('total_qty') or 0),2)
                    sheet.cell(row=total_row-6,column=month_data_col).alignment = right_aligment
                    sheet.cell(row=total_row-6,column=month_data_col).font = styles.Font(name='Arial',bold=False,size=12)
                    sheet.cell(row=total_row-6,column=month_data_col).fill = styles.PatternFill("solid",start_color="c3d69b")

                    sheet.cell(row=total_row-5,column=month_data_col).value = '$ {:,.2f}'.format(round((month_total_dict.get(y).get(month_data).get('total') or 0),2))
                    sheet.cell(row=total_row-5,column=month_data_col).alignment = right_aligment
                    sheet.cell(row=total_row-5,column=month_data_col).font = styles.Font(name='Arial',bold=False,size=12)
                    sheet.cell(row=total_row-5,column=month_data_col).fill = styles.PatternFill("solid",start_color="EBF1DE")

                    total_qty += (month_total_dict.get(y).get(month_data).get('total_qty') or 0)
                    main_total += (month_total_dict.get(y).get(month_data).get('total') or 0)
                    month_data_col += 1

                sheet.cell(row=total_row-7,column=month_data_col).value = 'Total'
                sheet.cell(row=total_row-7,column=month_data_col).alignment = center_aligment
                sheet.cell(row=total_row-7,column=month_data_col).fill = styles.PatternFill("solid",start_color="2f5496")
                sheet.cell(row=total_row-7,column=month_data_col).font = styles.Font(name='Arial',bold=True,color='ffffff',size=12)

                sheet.cell(row=total_row-6,column=month_data_col).value = round(total_qty,2)
                sheet.cell(row=total_row-6,column=month_data_col).alignment = right_aligment
                sheet.cell(row=total_row-6,column=month_data_col).fill = styles.PatternFill("solid",start_color="b4c6e7")

                sheet.cell(row=total_row-5,column=month_data_col).value = '$ {:,.2f}'.format(round(main_total,2))
                sheet.cell(row=total_row-5,column=month_data_col).alignment = right_aligment
                sheet.cell(row=total_row-5,column=month_data_col).fill = styles.PatternFill("solid",start_color="d9e2f3")

                month_data_col = content_col+1
                total_row += 3

            sheet.merge_cells('B%s:R%s'%(table_content_row+2,table_content_row+2))
            sheet.cell(row=table_content_row+2,column=2).fill = styles.PatternFill("solid",start_color="7f7f7f")
            sheet.row_dimensions[table_content_row+2].height = 10

            # Header of yearly total.
            merge_col_range = get_column_letter(20+len(year_total_dict)-1)
            sheet.merge_cells('T%s:%s%s'%(table_header_row+1,merge_col_range,table_header_row+1))
            sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).value = "YEARLY TOTALS"
            sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=False)
            sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).fill = styles.PatternFill("solid",start_color="2f5496")
            sheet.cell(row=table_header_row+1,column=year_table_col-len(year_range)).font = styles.Font(name='Arial',bold=True,color='ffffff')

            # Add value of yearly total.
            end_report_seperator = year_table_col-len(year_range)
            for r in year_range:
                year_total_data = year_total_dict.get(r)
                sheet.cell(row=table_content_row-7,column=end_report_seperator).value =round((year_total_data.get('total_qty') or 0.0),2)
                sheet.cell(row=table_content_row-7,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="b4c6e7")
                sheet.cell(row=table_content_row-6,column=end_report_seperator).value ='$ {:,.2f}'.format(round((year_total_data.get('total') or 0.0),2))
                sheet.cell(row=table_content_row-6,column=end_report_seperator).alignment = right_aligment
                sheet.cell(row=table_content_row-6,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="d9e2f3")
                sheet.cell(row=table_content_row+2,column=end_report_seperator).fill = styles.PatternFill("solid",start_color="7f7f7f")
                end_report_seperator += 1

        # Width of all col.
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 26
        sheet.column_dimensions['C'].width = 27
        sheet.column_dimensions['D'].width = 33
        sheet.column_dimensions['E'].width = 19
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 13
        sheet.column_dimensions['H'].width = 13
        sheet.column_dimensions['I'].width = 13
        sheet.column_dimensions['J'].width = 13
        sheet.column_dimensions['K'].width = 13
        sheet.column_dimensions['L'].width = 13
        sheet.column_dimensions['M'].width = 13
        sheet.column_dimensions['N'].width = 13
        sheet.column_dimensions['O'].width = 13
        sheet.column_dimensions['P'].width = 13
        sheet.column_dimensions['Q'].width = 13
        sheet.column_dimensions['R'].width = 13
        sheet.column_dimensions['S'].width = 2
        sheet.column_dimensions['T'].width = 16
        sheet.column_dimensions['U'].width = 16
        sheet.column_dimensions['V'].width = 16
        sheet.column_dimensions['W'].width = 16
        sheet.column_dimensions['X'].width = 16
        sheet.column_dimensions['Y'].width = 16
        sheet.column_dimensions['Z'].width = 16

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wizard_id = self.env['report.wizard'].create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wizard_id.id,f_name),
            'target': 'self',
        }



class ProductTemplate(models.Model):
    _inherit = 'product.template'

    is_kcash_rewards = fields.Boolean('K-Cash Product')
    hide_from_order = fields.Boolean('Hide From Order')

    